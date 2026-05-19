import json
import os
import shutil
from datetime import datetime, timedelta
from pathlib import Path

LOCAL_DIR = Path.home() / ".cambridge_tool"
CACHE_FILE = LOCAL_DIR / "cache.json"
CACHE_EXPIRY_DAYS = 7


class DataManager:
    def __init__(self, sync_dir: str = ""):
        LOCAL_DIR.mkdir(exist_ok=True)
        self._sync_dir: Path | None = self._resolve_sync_dir(sync_dir)
        if self._sync_dir:
            self._sync_dir.mkdir(parents=True, exist_ok=True)
        self.history   = self._load(self._history_file,   [])
        self.favorites = self._load(self._favorites_file, {})
        self._cache    = None   # lazy: loaded only on first word lookup

    @property
    def _history_file(self) -> Path:
        return (self._sync_dir / "history.json") if self._sync_dir \
               else (LOCAL_DIR / "history.json")

    @property
    def _favorites_file(self) -> Path:
        return (self._sync_dir / "favorites.json") if self._sync_dir \
               else (LOCAL_DIR / "favorites.json")

    def has_local_data(self) -> bool:
        return (LOCAL_DIR / "history.json").exists() \
            or (LOCAL_DIR / "favorites.json").exists()

    def sync_path_has_data(self, path: str) -> bool:
        p = self._resolve_sync_dir(path)
        return p is not None and (
            (p / "history.json").exists() or (p / "favorites.json").exists()
        )

    def migrate_local_to_path(self, dest: str):
        """Copy local history/favorites to dest/HotDict/ (skip if already present)."""
        dest_dir = self._resolve_sync_dir(dest)
        dest_dir.mkdir(parents=True, exist_ok=True)
        for fname in ("history.json", "favorites.json"):
            src = LOCAL_DIR / fname
            dst = dest_dir / fname
            if src.exists() and not dst.exists():
                shutil.copy2(src, dst)

    @staticmethod
    def _resolve_sync_dir(path: str) -> Path | None:
        """Append HotDict subfolder so files don't clutter the cloud root."""
        return (Path(path) / "HotDict") if path else None

    def set_sync_dir(self, path: str):
        """Switch active data directory and reload history/favorites."""
        self._sync_dir = self._resolve_sync_dir(path)
        if self._sync_dir:
            self._sync_dir.mkdir(parents=True, exist_ok=True)
        self.history   = self._load(self._history_file,   [])
        self.favorites = self._load(self._favorites_file, {})

    @property
    def cache(self):
        if self._cache is None:
            self._cache = self._load(CACHE_FILE, {})
        return self._cache

    @cache.setter
    def cache(self, value):
        self._cache = value

    def _load(self, path, default):
        try:
            if path.exists():
                with open(path, encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return default

    def _save(self, path, data):
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Save error: {e}")

    # ── Cache ────────────────────────────────────────────────────────────────

    def get_cached(self, word: str):
        key = word.lower().strip()
        entry = self.cache.get(key)
        if entry:
            try:
                ts = datetime.fromisoformat(entry["cached_at"])
                if datetime.now() - ts < timedelta(days=CACHE_EXPIRY_DAYS):
                    return entry["data"]
            except Exception:
                pass
            del self.cache[key]
        return None

    def set_cached(self, word: str, data: dict):
        key = word.lower().strip()
        self.cache[key] = {"cached_at": datetime.now().isoformat(), "data": data}
        self._save(CACHE_FILE, self.cache)

    # ── History ──────────────────────────────────────────────────────────────

    def add_history(self, word: str):
        word = word.strip()
        if not word:
            return
        self.history = [h for h in self.history if h["word"].lower() != word.lower()]
        self.history.insert(0, {"word": word, "time": datetime.now().isoformat()})
        self.history = self.history[:200]
        self._save(self._history_file, self.history)

    def get_history(self) -> list:
        return [h["word"] for h in self.history]

    def get_today_history_count(self) -> int:
        today = datetime.now().date()
        count = 0
        for h in self.history:
            try:
                if datetime.fromisoformat(h["time"]).date() == today:
                    count += 1
            except Exception:
                pass
        return count

    def remove_cached(self, word: str):
        key = word.lower().strip()
        if key in self.cache:
            del self.cache[key]
            self._save(CACHE_FILE, self.cache)

    def clear_cache(self):
        self.cache = {}
        self._save(CACHE_FILE, self.cache)

    def remove_history(self, word: str):
        self.history = [h for h in self.history if h["word"].lower() != word.lower()]
        self._save(self._history_file, self.history)

    def clear_history(self):
        self.history = []
        self._save(self._history_file, self.history)
        self.clear_cache()

    # ── Favorites ────────────────────────────────────────────────────────────

    def toggle_favorite(self, word: str, data: dict = None) -> bool:
        """Returns True if now a favorite, False if removed."""
        key = word.lower().strip()
        if key in self.favorites:
            del self.favorites[key]
            self._save(self._favorites_file, self.favorites)
            return False
        self.favorites[key] = {
            "word": word.strip(),
            "data": data or {},
            "time": datetime.now().isoformat(),
        }
        self._save(self._favorites_file, self.favorites)
        return True

    def is_favorite(self, word: str) -> bool:
        return word.lower().strip() in self.favorites

    def get_favorites(self) -> list:
        return [v["word"] for v in self.favorites.values()]

    def remove_favorite(self, word: str):
        key = word.lower().strip()
        if key in self.favorites:
            del self.favorites[key]
            self._save(self._favorites_file, self.favorites)

    def clear_favorites(self):
        self.favorites = {}
        self._save(self._favorites_file, self.favorites)
        self.clear_cache()

    def update_favorite_data(self, word: str, data: dict):
        key = word.lower().strip()
        if key in self.favorites:
            self.favorites[key]["data"] = data
            self._save(self._favorites_file, self.favorites)

    # ── Export ───────────────────────────────────────────────────────────────

    def export_favorites_xlsx(self, path: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "收藏单词"

        headers = ["单词", "音标", "词性", "英文释义", "中文释义", "收藏时间"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 16

        row = 2
        for fav in self.favorites.values():
            word = fav["word"]
            data = fav.get("data") or {}
            time_str = fav.get("time", "")[:10]

            prons = data.get("pronunciations", [])
            pron_str = "  ".join(
                f"{'uk' if 'uk' in p.get('label','').lower() else 'us' if 'us' in p.get('label','').lower() else p.get('label','')} /{p['ipa']}/"
                for p in prons if p.get("ipa")
            )

            entries = data.get("entries", [])
            if not entries:
                ws.append([word, pron_str, "", "", "", time_str])
                row += 1
                continue

            first = True
            for entry in entries:
                pos = entry.get("pos", "")
                for defn in entry.get("definitions", []):
                    ws.cell(row=row, column=1, value=word if first else "")
                    ws.cell(row=row, column=2, value=pron_str if first else "")
                    ws.cell(row=row, column=3, value=pos)
                    ws.cell(row=row, column=4, value=defn.get("en", ""))
                    ws.cell(row=row, column=5, value=defn.get("zh", ""))
                    ws.cell(row=row, column=6, value=time_str if first else "")
                    row += 1
                    first = False

        wb.save(path)
